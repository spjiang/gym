"""商户会员 Excel 导入。"""

from io import BytesIO

from fastapi.testclient import TestClient
from openpyxl import Workbook, load_workbook

from app.core.errors import AppError
from app.systems.platform.services.member_import import (
    build_member_import_template,
    normalize_phone,
    parse_member_import,
)


def _xlsx(rows: list[tuple], headers: tuple[str, str] = ("手机号", "姓名")) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "会员名单"
    ws.append(list(headers))
    for row in rows:
        ws.append(list(row))
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def _merchant_id(client: TestClient, headers: dict, name_part: str = "FIT") -> int:
    merchants = client.get("/api/v1/merchants", headers=headers).json()
    hit = next((m for m in merchants if name_part in m["name"]), merchants[0])
    return hit["id"]


def test_normalize_phone():
    assert normalize_phone("138-0013-8000") == "13800138000"
    assert normalize_phone("86 13900001111") == "13900001111"
    assert normalize_phone("1380013800") is None
    assert normalize_phone("23800138000") is None


def test_parse_skips_example_and_reports_invalid():
    raw = _xlsx(
        [
            ("13800138000", "示例会员（请删除本行后填写真实数据）"),
            ("13900001111", "张三"),
            ("bad", "李四"),
            ("13900001111", "重复"),
            ("", ""),
        ]
    )
    rows, issues = parse_member_import(raw)
    assert [r.phone for r in rows] == ["13900001111"]
    assert any("11 位" in i.message for i in issues)
    assert any("重复" in i.message for i in issues)


def test_download_import_template(client: TestClient, admin_headers: dict):
    resp = client.get("/api/v1/members/import-template", headers=admin_headers)
    assert resp.status_code == 200, resp.text
    assert resp.content[:2] == b"PK"
    wb = load_workbook(BytesIO(resp.content))
    assert "会员名单" in wb.sheetnames
    assert [c.value for c in wb["会员名单"][1]][:2] == ["手机号", "姓名"]


def test_site_admin_imports_and_links_existing(client: TestClient, admin_headers: dict):
    merchants = client.get("/api/v1/merchants", headers=admin_headers).json()
    gym_id = merchants[0]["id"]
    bar_id = merchants[1]["id"] if len(merchants) > 1 else gym_id

    created = client.post(
        "/api/v1/members",
        headers=admin_headers,
        json={"phone": "13900002222", "name": "已有会员", "merchant_id": gym_id},
    )
    assert created.status_code == 200, created.text

    payload = _xlsx(
        [
            ("13900003333", "新会员甲"),
            ("13900002222", "不会覆盖"),
            ("12345", "坏号"),
        ]
    )
    first = client.post(
        "/api/v1/members/import",
        headers=admin_headers,
        params={"merchant_id": gym_id},
        files={"file": ("members.xlsx", payload, "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")},
    )
    assert first.status_code == 200, first.text
    body = first.json()
    assert body["created"] == 1
    assert body["linked"] == 0
    assert body["skipped"] == 1
    assert body["failed"] == 1
    assert body["merchant_id"] == gym_id

    listed = client.get("/api/v1/members", headers=admin_headers, params={"q": "13900002222"}).json()
    assert listed["items"][0]["name"] == "已有会员"

    if bar_id != gym_id:
        second = client.post(
            "/api/v1/members/import",
            headers=admin_headers,
            params={"merchant_id": bar_id},
            files={
                "file": (
                    "members.xlsx",
                    _xlsx([("13900003333", "新会员甲")]),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert second.status_code == 200, second.text
        assert second.json()["linked"] == 1
        assert second.json()["created"] == 0
        member = client.get("/api/v1/members", headers=admin_headers, params={"q": "13900003333"}).json()["items"][0]
        assert gym_id in member["merchant_ids"]
        assert bar_id in member["merchant_ids"]
        assert member["first_merchant_id"] == gym_id


def test_import_requires_merchant_for_site_admin(client: TestClient, admin_headers: dict):
    resp = client.post(
        "/api/v1/members/import",
        headers=admin_headers,
        files={
            "file": (
                "members.xlsx",
                _xlsx([("13900004444", "丙")]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert resp.status_code == 400


def test_merchant_admin_cannot_import_other_shop(client: TestClient, admin_headers: dict):
    merchants = client.get("/api/v1/merchants", headers=admin_headers).json()
    gym_id = merchants[0]["id"]
    other_id = merchants[1]["id"] if len(merchants) > 1 else None
    staff = client.post(
        "/api/v1/staff",
        headers=admin_headers,
        json={
            "username": "import_admin",
            "password": "Merchant@1",
            "display_name": "导入管理员",
            "merchant_id": gym_id,
            "role_codes": ["gym_admin"],
        },
    )
    assert staff.status_code == 200, staff.text
    token = client.post(
        "/api/v1/auth/login", json={"username": "import_admin", "password": "Merchant@1"}
    ).json()["access_token"]
    headers = {"Authorization": f"Bearer {token}"}

    own = client.post(
        "/api/v1/members/import",
        headers=headers,
        files={
            "file": (
                "members.xlsx",
                _xlsx([("13900005555", "本店会员")]),
                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            )
        },
    )
    assert own.status_code == 200, own.text
    assert own.json()["created"] == 1
    assert own.json()["merchant_id"] == gym_id

    if other_id is not None:
        denied = client.post(
            "/api/v1/members/import",
            headers=headers,
            params={"merchant_id": other_id},
            files={
                "file": (
                    "members.xlsx",
                    _xlsx([("13900006666", "跨店")]),
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
        )
        assert denied.status_code == 403


def test_reject_non_xlsx(client: TestClient, admin_headers: dict):
    mid = _merchant_id(client, admin_headers)
    resp = client.post(
        "/api/v1/members/import",
        headers=admin_headers,
        params={"merchant_id": mid},
        files={"file": ("members.csv", b"phone,name\n13900007777,Ding", "text/csv")},
    )
    assert resp.status_code == 400


def test_template_bytes_are_parseable():
    data = build_member_import_template()
    try:
        parse_member_import(data)
        raise AssertionError("仅含示例行时应拒绝")
    except AppError as exc:
        assert "没有可导入" in exc.message
