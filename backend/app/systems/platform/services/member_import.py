"""商户会员 Excel 导入：模板、解析与挂靠。"""

from __future__ import annotations

import re
from dataclasses import dataclass
from io import BytesIO

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.datavalidation import DataValidation
from sqlalchemy import select
from sqlalchemy.orm import Session

from app.core.errors import AppError
from app.systems.platform.models.member import AcquisitionSource, FaceStatus, Member, MerchantMember
from app.systems.platform.services.promotion import ensure_member_promoter_code

PHONE_RE = re.compile(r"^1[3-9]\d{9}$")
PHONE_HEADERS = {"手机号", "手机", "电话", "会员手机", "phone", "mobile"}
NAME_HEADERS = {"姓名", "会员姓名", "名称", "name"}
MAX_IMPORT_ROWS = 2000
MAX_IMPORT_BYTES = 2 * 1024 * 1024
XLSX_MIME = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"


@dataclass(frozen=True)
class ImportRow:
    row: int
    phone: str
    name: str


@dataclass(frozen=True)
class ImportIssue:
    row: int
    phone: str | None
    name: str | None
    message: str


def normalize_phone(raw: str) -> str | None:
    """去掉空格与分隔符，校验为中国大陆 11 位手机号。"""
    text = re.sub(r"[\s\-()（）+]", "", (raw or "").strip())
    if text.startswith("86") and len(text) == 13:
        text = text[2:]
    if text.endswith(".0") and text[:-2].isdigit():
        text = text[:-2]
    return text if PHONE_RE.fullmatch(text) else None


def cell_text(value: object) -> str:
    if value is None:
        return ""
    if isinstance(value, bool):
        return ""
    if isinstance(value, float) and value.is_integer():
        return str(int(value))
    if isinstance(value, int):
        return str(value)
    return str(value).strip()


def build_member_import_template() -> bytes:
    """生成带说明与表头的 xlsx 模板。"""
    wb = Workbook()
    guide = wb.active
    guide.title = "导入说明"
    guide["A1"] = "商户会员导入说明"
    guide["A1"].font = Font(bold=True, size=14)
    lines = [
        "1. 请在「会员名单」工作表从第 2 行起填写，不要改动表头。",
        "2. 必填列：手机号、姓名。手机号须为 11 位中国大陆号码（1 开头）。",
        "3. 导入后会员将挂靠到当前所选商户；同一场地内手机号唯一。",
        "4. 若手机号已在本场地存在，将只补挂靠关系，不会覆盖原姓名。",
        "5. 已挂靠本店的会员会跳过。单次最多 2000 行，仅支持 .xlsx。",
        "6. 请勿在表格中填写密码或证件号。",
    ]
    for index, text in enumerate(lines, start=3):
        guide[f"A{index}"] = text
        guide[f"A{index}"].alignment = Alignment(wrap_text=True)
    guide.column_dimensions["A"].width = 88

    sheet = wb.create_sheet("会员名单", 0)
    headers = ["手机号", "姓名"]
    header_fill = PatternFill("solid", fgColor="171B1F")
    header_font = Font(bold=True, color="F2E6D2")
    for col, title in enumerate(headers, start=1):
        cell = sheet.cell(1, col, title)
        cell.fill = header_fill
        cell.font = header_font
        cell.alignment = Alignment(horizontal="center")
        sheet.column_dimensions[get_column_letter(col)].width = 22
    sheet.cell(2, 1, "13800138000")
    sheet.cell(2, 2, "示例会员（请删除本行后填写真实数据）")
    sheet.freeze_panes = "A2"
    phone_dv = DataValidation(type="textLength", operator="equal", formula1="11", allow_blank=True)
    phone_dv.error = "请填写 11 位手机号"
    phone_dv.errorTitle = "手机号格式"
    phone_dv.add("A2:A2001")
    sheet.add_data_validation(phone_dv)

    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def parse_member_import(data: bytes) -> tuple[list[ImportRow], list[ImportIssue]]:
    """解析 xlsx，返回合法行与行级错误。"""
    if not data:
        raise AppError("invalid_file", "文件为空", status_code=400)
    if len(data) > MAX_IMPORT_BYTES:
        raise AppError("invalid_file", "导入文件不能超过 2MB", status_code=400)
    if data[:2] != b"PK":
        raise AppError("invalid_file", "仅支持 .xlsx 模板文件", status_code=400)

    try:
        wb = load_workbook(BytesIO(data), read_only=True, data_only=True)
    except Exception as exc:
        raise AppError("invalid_file", "无法读取 Excel，请使用系统提供的模板", status_code=400) from exc

    sheet = wb["会员名单"] if "会员名单" in wb.sheetnames else wb[wb.sheetnames[0]]
    rows = list(sheet.iter_rows(values_only=True))
    wb.close()
    if not rows:
        raise AppError("invalid_file", "表格为空", status_code=400)

    header_index = _find_header_row(rows)
    if header_index is None:
        raise AppError("invalid_file", "未找到表头，请保留「手机号」「姓名」两列", status_code=400)

    phone_col, name_col = _header_columns(rows[header_index])
    parsed: list[ImportRow] = []
    issues: list[ImportIssue] = []
    seen: dict[str, int] = {}
    data_count = 0

    for offset, raw in enumerate(rows[header_index + 1 :], start=header_index + 2):
        phone_raw = cell_text(_col(raw, phone_col))
        name_raw = cell_text(_col(raw, name_col))
        if not phone_raw and not name_raw:
            continue
        if "示例" in name_raw:
            continue
        data_count += 1
        if data_count > MAX_IMPORT_ROWS:
            issues.append(ImportIssue(offset, phone_raw or None, name_raw or None, f"超过 {MAX_IMPORT_ROWS} 行上限"))
            break
        if not phone_raw or not name_raw:
            issues.append(ImportIssue(offset, phone_raw or None, name_raw or None, "手机号与姓名均为必填"))
            continue
        if len(name_raw) > 128:
            issues.append(ImportIssue(offset, phone_raw, name_raw, "姓名不能超过 128 字"))
            continue
        phone = normalize_phone(phone_raw)
        if phone is None:
            issues.append(ImportIssue(offset, phone_raw, name_raw, "手机号须为 11 位中国大陆号码"))
            continue
        if phone in seen:
            issues.append(ImportIssue(offset, phone, name_raw, f"与第 {seen[phone]} 行手机号重复"))
            continue
        seen[phone] = offset
        parsed.append(ImportRow(row=offset, phone=phone, name=name_raw))

    if not parsed and not issues:
        raise AppError("invalid_file", "没有可导入的数据行", status_code=400)
    return parsed, issues


def import_members_for_merchant(
    db: Session,
    *,
    site_id: int,
    merchant_id: int,
    rows: list[ImportRow],
) -> dict[str, int]:
    """按手机号创建或挂靠会员，返回计数。"""
    created = linked = skipped = 0
    phones = [row.phone for row in rows]
    existing = {
        m.phone: m
        for m in db.scalars(select(Member).where(Member.site_id == site_id, Member.phone.in_(phones))).all()
    }
    member_ids = [m.id for m in existing.values()]
    already_linked: set[int] = set()
    if member_ids:
        already_linked = set(
            db.scalars(
                select(MerchantMember.member_id).where(
                    MerchantMember.merchant_id == merchant_id,
                    MerchantMember.member_id.in_(member_ids),
                )
            ).all()
        )

    for row in rows:
        member = existing.get(row.phone)
        if member is None:
            member = Member(
                site_id=site_id,
                phone=row.phone,
                name=row.name,
                face_status=FaceStatus.NOT_ENROLLED.value,
                acquisition_source=AcquisitionSource.MERCHANT.value,
                first_merchant_id=merchant_id,
            )
            db.add(member)
            db.flush()
            ensure_member_promoter_code(db, member)
            existing[row.phone] = member
            db.add(MerchantMember(merchant_id=merchant_id, member_id=member.id))
            created += 1
            continue
        if member.id in already_linked:
            skipped += 1
            continue
        db.add(MerchantMember(merchant_id=merchant_id, member_id=member.id))
        already_linked.add(member.id)
        linked += 1
    return {"created": created, "linked": linked, "skipped": skipped}


def _find_header_row(rows: list[tuple]) -> int | None:
    for index, raw in enumerate(rows[:8]):
        labels = {cell_text(v).lower() for v in raw if cell_text(v)}
        if labels & {h.lower() for h in PHONE_HEADERS} and labels & {h.lower() for h in NAME_HEADERS}:
            return index
    return None


def _header_columns(header: tuple) -> tuple[int, int]:
    phone_col = name_col = None
    for index, value in enumerate(header):
        label = cell_text(value).lower()
        if label in {h.lower() for h in PHONE_HEADERS}:
            phone_col = index
        elif label in {h.lower() for h in NAME_HEADERS}:
            name_col = index
    if phone_col is None or name_col is None:
        raise AppError("invalid_file", "表头需同时包含「手机号」和「姓名」", status_code=400)
    return phone_col, name_col


def _col(raw: tuple, index: int) -> object:
    if index >= len(raw):
        return None
    return raw[index]
